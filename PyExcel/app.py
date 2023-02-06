import openpyxl

# # CREATE WORKBOOK
# # wb = openpyxl.Workbook()

# # LOADING WORKBOOK
# wb = openpyxl.load_workbook("transaction.xlsx")
# print(wb.sheetnames)  # ['Sheet1']

# # ACCESS SHEET
# sheet = wb["Sheet1"]

# # CREATE NEW SHEET
# # wb.create_sheet("Sheet2", 0) # second is index , which means index 0

# # REMOVE NEW SHEET
# # wb.remove_sheet(sheet)

# # ACCESS CELL
# cell = sheet["A1"]
# # cell = sheet.cell(row=1,column=1) # ANOTHER WAY

# print(cell.value)  # transaction_id

# # modify the value of cell
# cell.value = "transactions_id"

# # ACCESS ROW AND COL
# print(cell.row)  # 1
# print(cell.column)  # 1
# print(cell.coordinate)  # A1
# print(sheet.max_row)  # 4
# print(sheet.max_column)  # 3

# # ACESSS ALL VALUE OF SHEET
# for row in range(1, sheet.max_row+1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# # transactions_id
# # product_id
# # price
# # 1001
# # 1
# # 5.95
# # 1002
# # 2
# # =C2+1
# # 1003
# # 3
# # =C3+1

# # ACESS RANGE OF CELL
# columna = sheet["a"]
# print(columna)
# # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>)(PyExcel)

# print(sheet["a:c"])
# # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>), (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>), (<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.C4>))

# print(sheet["a1:c3"])
# # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

# print(sheet["1:3"])
# # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
# # (PyExcel)

# # ADD ROW TO SHEET
# sheet.append([1, 2, 3])

# # INSERT ROW AT GIVEN INDEX
# # sheet.insert_rows()
# # sheet.insert_cols()
# # sheet.delete_cols()
# # sheet.delete_cols()

# wb.save("transaction2.xlsx")
# # DOCUMENTATION
# # https: // openpyxl.readthedocs.io/en/stable/tutorial.html

# COMMAND QUERY SEPARATION PRINCIPLE
# state a method or function should either perform an action and change the state of system
# or query that returned anwser to the system without change the system and have side effect but not both

wb = openpyxl.load_workbook("transaction.xlsx")
# command method - responsible create the data in the system
# wb.create_sheet()
sheet = wb["Sheet1"]
# query method  -but violate the principle, because it will also creat the data in the system
# shell.cell()
# better to raise exception , if access cell out of range
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)

sheet.append([1, 2, 3])
# transaction_id
# 1001
# 1002
# 1003
# None
# None
# None
# None
# None

wb.save("transactions3.xlsx")

# IN EXCEL - when we loop the sheet, it will create the new cell None value to the cell that out of range
# transaction_id	product_id	price
# 1001	              1	         $5.95
# 1002	              2	         $6.95
# 1003	              3	         $7.95


# 1	                   2	        3


# numbers = [1,2,3]
#RAISE EXCEPTION
# numbers[4]
# dic['invalidkey']